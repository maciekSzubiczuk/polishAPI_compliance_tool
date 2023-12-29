from flask import Flask, render_template, request, redirect, url_for, jsonify, make_response,send_file
import yaml
from deepdiff import DeepDiff
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Border, Side

app = Flask(__name__)

# Storage for uploaded file contents
polish_api_data = None
santander_api_data = []
# Temporary storage for the CSV file
xlsx_file_path = 'C:\praca_inzynierska\polishAPI_compliance_tool\\reports\\report.xlsx'  # Adjust the path as needed

api_sections = {
    'PIS': '/v2_1_1.1/payments',
    'CAF': '/v2_1_1.1/confirmation',
    'AIS': '/v2_1_1.1/accounts',
    'AS': '/v2_1_1.1/auth'
}

def find_definitions_differences(polish_api_data, santander_api_data):
    polish_definitions = polish_api_data.get('definitions', {})
    santander_definitions = santander_api_data.get('definitions', {})
    return find_differences(polish_definitions, santander_definitions)

def categorize_paths(paths, api_sections):
    categorized_paths = {section: {} for section in api_sections}
    for path, details in paths.items():
        for section, pattern in api_sections.items():
            if path.startswith(pattern):
                categorized_paths[section][path] = details
                break
    return categorized_paths

def find_differences_by_section(polish_api_data, santander_api_data, api_sections):
    differences_by_section = {section: {} for section in api_sections}

    polish_categorized = categorize_paths(polish_api_data.get('paths', {}), api_sections)
    santander_categorized = categorize_paths(santander_api_data.get('paths', {}), api_sections)

    for section in api_sections:
        differences_by_section[section] = find_differences(
            polish_categorized[section], 
            santander_categorized[section]
        )

    return differences_by_section



def merge_api_data(api_files):
    merged_data = {}
    for api_file in api_files:
        api_data = load_yaml_from_file(api_file)
        for key, value in api_data.items():
            if key in merged_data:
                if isinstance(merged_data[key], dict) and isinstance(value, dict):
                    merged_data[key].update(value)
                elif isinstance(merged_data[key], list) and isinstance(value, list):
                    # Extend the list with unique items
                    for item in value:
                        if item not in merged_data[key]:
                            merged_data[key].append(item)
                else:
                    pass
            else:
                merged_data[key] = value
    return merged_data

def load_yaml_from_file(file):
    if file:
        return yaml.safe_load(file)

def generate_summary(change):
    summary = []

    def format_yaml(value):
        if isinstance(value, (dict, list)):
            return yaml.dump(value, default_flow_style=False, sort_keys=False).strip()
        return value

    def format_change(path, value, change_type):
        if isinstance(value, dict):
            # Format each key-value pair in the dictionary with a newline and a dash
            formatted_value = "\n".join([f"- {k}: {format_yaml(v)}" for k, v in value.items()])
        else:
            formatted_value = f"- {format_yaml(value)}"

        if path == 'Root':  # Avoid including 'Root' in the path
            return f"{change_type}:\n{formatted_value}"
        else:
            return f"{change_type}:\n{path}\n{formatted_value}"

    def compare_dicts(left, right, path=''):
        keys = set(left.keys()).union(right.keys())
        for key in keys:
            left_value = left.get(key)
            right_value = right.get(key)
            new_path = f"{path}.{key}" if path else key

            if left_value is None:
                summary.append(format_change(new_path, right_value, "Added"))
            elif right_value is None:
                summary.append(format_change(new_path, left_value, "Removed"))
            elif isinstance(left_value, dict) and isinstance(right_value, dict):
                compare_dicts(left_value, right_value, new_path)
            elif isinstance(left_value, list) and isinstance(right_value, list):
                compare_lists(left_value, right_value, new_path)
            elif left_value != right_value:
                if right_value is None:
                    summary.append(format_change(new_path, left_value, "Removed"))
                else:
                    summary.append(format_change(new_path, right_value, "Added"))

    def compare_lists(left, right, path):
        added_items = [item for item in right if item not in left]
        removed_items = [item for item in left if item not in right]

        for item in added_items:
            summary.append(format_change(path, item, "Added"))
        for item in removed_items:
            summary.append(format_change(path, item, "Removed"))

    # Handling of root level differences
    if isinstance(change['left'], dict) and isinstance(change['right'], dict):
        compare_dicts(change['left'], change['right'])
    elif isinstance(change['left'], list) and isinstance(change['right'], list):
        compare_lists(change['left'], change['right'], '')
    else:
        # Handling simple types like strings at the root level
        compare_dicts({'Root': change['left']}, {'Root': change['right']}, '')

    final_summary = '\n'.join(summary)
    return final_summary

def find_differences(dict1, dict2, base_path=''):
    differences = {}
    for key in dict1:
        if key not in dict2:
            differences[base_path + key] = {'left': dict1[key], 'right': None}
        elif dict1[key] != dict2[key]:
            if isinstance(dict1[key], dict) and isinstance(dict2[key], dict):
                sub_diffs = find_differences(dict1[key], dict2[key], base_path + key + '.')
                differences.update(sub_diffs)
            else:
                differences[base_path + key] = {'left': dict1[key], 'right': dict2[key]}
    for key in dict2:
        if key not in dict1:
            differences[base_path + key] = {'left': None, 'right': dict2[key]}
    return differences

@app.route('/', methods=['GET', 'POST'])
def index():
    global polish_api_data, santander_api_data
    if request.method == 'POST':
        polish_api_file = request.files.get('polishapi_file')
        santander_files = request.files.getlist('santander_files')

        if polish_api_file:
            polish_api_data = load_yaml_from_file(polish_api_file)

        if santander_files:
            santander_api_data = merge_api_data(santander_files)

        return redirect(url_for('display'))

    return render_template('upload.html')


@app.route('/display')
def display():
    if not polish_api_data or not santander_api_data:
        return render_template('display.html', differences_by_section={})

    # Find differences by API section
    differences_by_section = find_differences_by_section(polish_api_data, santander_api_data, api_sections)

    # Find differences in definitions
    definitions_differences = find_definitions_differences(polish_api_data, santander_api_data)
    
    # Combine all differences and generate summaries
    all_differences = {}
    for section, diffs in differences_by_section.items():
        formatted_diffs = {}
        for path, change in diffs.items():
            summary = generate_summary(change)
            formatted_diffs[path] = {
                'left': yaml.dump(change['left'], default_flow_style=False, sort_keys=False) if change['left'] else '',
                'right': yaml.dump(change['right'], default_flow_style=False, sort_keys=False) if change['right'] else '',
                'summary': summary
            }
        all_differences[section] = formatted_diffs

    # Add definitions differences with summary
    formatted_definitions_diffs = {}
    for key, change in definitions_differences.items():
        summary = generate_summary(change)
        formatted_definitions_diffs[key] = {
            'left': yaml.dump(change['left'], default_flow_style=False, sort_keys=False) if change['left'] else '',
            'right': yaml.dump(change['right'], default_flow_style=False, sort_keys=False) if change['right'] else '',
            'summary': summary
        }
    all_differences['Definitions'] = formatted_definitions_diffs

    wb = Workbook()
    ws = wb.active

    # Update column headers
    headers = ['Section', 'Path', 'PolishApi', 'Santander', 'Summary']
    ws.append(headers)

    # Add filters to the top of each column
    ws.auto_filter.ref = ws.dimensions

    # Write the data
    for section, diffs in all_differences.items():
        for path, change in diffs.items():
            ws.append([
                section,
                path,
                change.get('left', ''),
                change.get('right', ''),
                change.get('summary', '')
            ])

    # Change header colors
    header_colors = {
        'A': 'D3D3D3',  # Light gray
        'B': 'DDA0DD',  # Light purple
        'C': '90EE90',  # Light green
        'D': 'FF6347',  # Darker red for Santander
        'E': 'F0F0F0',  # Very light gray for Summary
    }

    # Apply header colors
    for column, color in header_colors.items():
        ws[column + '1'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # More subtle colors for sections
    section_colors = {
        'PIS': 'FFF0E0', 'CAF': 'E0FFF0', 'AIS': 'E0E0FF', 'AS': 'FFE0E0', 'Definitions': 'FFF0FF'
    }
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
        section = row[0].value
        if section in section_colors:
            row[0].fill = PatternFill(start_color=section_colors[section], end_color=section_colors[section], fill_type="solid")

    # Apply borders for better visibility
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Adjust Column Widths
    column_widths = {'B': 60, 'C': 45, 'D': 35, 'E': 35}  # Increased widths
    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width

    # Delete unnecessary column 'F'
    ws.delete_cols(6)

    # Additional Improvements: Center Alignment
    top_left_alignment = Alignment(horizontal='left', vertical='top',wrap_text=True)
    # Adjust Row Heights to fit content
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = top_left_alignment
            new_height = max(cell.value.count('\n') + 1, 1) * 15  # Adjust height based on content
            if ws.row_dimensions[cell.row].height is None or \
               ws.row_dimensions[cell.row].height < new_height:
                ws.row_dimensions[cell.row].height = new_height

    ws.freeze_panes = 'A2'

    # Save the workbook to a temporary file
    xlsx_file_path_temp = xlsx_file_path
    wb.save(xlsx_file_path_temp)

    # Calculate counts for additions and deletions
    counts_by_section = {}
    for section, diffs in all_differences.items():
        additions = sum(1 for change in diffs.values() if not change.get('left'))
        deletions = sum(1 for change in diffs.values() if not change.get('right'))
        counts_by_section[section] = {'additions': additions, 'deletions': deletions}

    # Pass these counts along with all_differences to the template
    return render_template('display.html', differences_by_section=all_differences, counts_by_section=counts_by_section)

@app.route('/download-xlsx')
def download_xlsx():
    return send_file(xlsx_file_path, as_attachment=True, download_name='report.xlsx')

if __name__ == '__main__':
    app.run(debug=True)


