from openpyxl import Workbook
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
import yaml
import openpyxl.styles as styles  

def generate_summary(change):
    summary = []

    def format_yaml(value, level=0, is_top_level=False):
        indent = "  " * level
        if isinstance(value, dict):
            formatted_items = []
            for k, v in value.items():
                prefix = "• " if is_top_level else ""
                if isinstance(v, (dict, list)):
                    formatted_items.append(f"{indent}{prefix}{k}:\n{format_yaml(v, level + 1)}")
                else:
                    formatted_items.append(f"{indent}{prefix}{k}: {v}")
            return "\n".join(formatted_items)
        elif isinstance(value, list):
            formatted_list = []
            for v in value:
                prefix = "• " if is_top_level else "- "
                formatted_list.append(f"{indent}{prefix}{format_yaml(v, level + 1, False)}")
            return "\n".join(formatted_list)
        else:
            return f"{indent}{'• ' if is_top_level else ''}{value}"
        
    def format_change(path, value, change_type):
        if value is None or (isinstance(value, (dict, list)) and not value):
            return ''
        formatted_value = format_yaml(value, 0, is_top_level=True)
        prefix = f"{change_type}:" if path == 'Root' else f"{change_type}:\n{path}"
        return f"{prefix}\n{formatted_value}"

    def compare_dicts(left, right, path=''):
        keys = set(left.keys()).union(right.keys())
        for key in keys:
            left_value = left.get(key)
            right_value = right.get(key)
            new_path = f"{path}.{key}" if path else key

            if left_value is None and right_value:
                formatted_change = format_change(new_path, right_value, "Added")
                if formatted_change.strip():
                    summary.append(formatted_change)
            elif right_value is None and left_value:
                formatted_change = format_change(new_path, left_value, "Removed")
                if formatted_change.strip():
                    summary.append(formatted_change)
            elif isinstance(left_value, dict) and isinstance(right_value, dict):
                compare_dicts(left_value, right_value, new_path)
            elif isinstance(left_value, list) and isinstance(right_value, list):
                compare_lists(left_value, right_value, new_path)
            elif left_value != right_value:
                if right_value is None or left_value is None:
                    pass
                else:
                    formatted_change = format_change(new_path, right_value if left_value is None else left_value, "Removed" if right_value is None else "Added")
                    if formatted_change.strip():
                        summary.append(formatted_change)


    def compare_lists(left, right, path):
        added_items = [item for item in right if item not in left]
        removed_items = [item for item in left if item not in right]

        for item in added_items:
            summary.append(format_change(path, item, "Added"))
        for item in removed_items:
            summary.append(format_change(path, item, "Removed"))

    if isinstance(change['left'], dict) and isinstance(change['right'], dict):
        compare_dicts(change['left'], change['right'])
    elif isinstance(change['left'], list) and isinstance(change['right'], list):
        compare_lists(change['left'], change['right'], '')
    else:
        compare_dicts({'Root': change['left']}, {'Root': change['right']}, '')

    final_summary = '\n'.join(summary)
    return final_summary



def generate_excel_report(formatted_diffs, xlsx_file_path):
  """
  Generates an Excel report from a list of diffs.

  Args:
      formatted_diffs: A list of dictionaries containing diff information.
      xlsx_file_path: The path to save the Excel report.
  """
  wb = openpyxl.Workbook() 
  ws = wb.active

  headers = ['Section', 'Status', 'Path', 'PolishApi', 'Santander', 'Summary']
  ws.append(headers)
  ws.auto_filter.ref = ws.dimensions

  for diff in formatted_diffs:
    ws.append([
      diff['section'],
      diff['status'],
      diff['path'],
      diff['left'],
      diff['right'],
      diff['summary']
    ])

  header_colors = {
  'A': 'D3D3D3',  # LIGHT GREY
  'B': 'FFCC99',  # LIGHT ORANGE
  'C': 'DDA0DD',  # LIGHT PURPLE
  'D': '90EE90',  # LIGHT GREEN
  'E': 'FFC0CB',  # LIGHT RED
  'F': 'C6E0FF',  # LIGHT BLUE
}

  for column, color in header_colors.items():
    ws[column + '1'].fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type="solid")

  summary_column = ws['F'] 
  for cell in summary_column:
    if cell.value: 
      cell.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='top', wrap_text=True)

  section_colors = {
    'PIS': 'FFF0E0', 'CAF': 'E0FFF0', 'AIS': 'E0E0FF', 'AS': 'FFE0E0', 'Definitions': 'FFF0FF'
  }
  for row in ws.iter_rows(min_row=2, max_col=1, values_only=False):
    section = row[0].value
    if section in section_colors:
      row[0].fill = openpyxl.styles.PatternFill(start_color=section_colors[section], end_color=section_colors[section], fill_type="solid")

  thin_border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                       right=openpyxl.styles.Side(style='thin'),
                                       top=openpyxl.styles.Side(style='thin'),
                                       bottom=openpyxl.styles.Side(style='thin'))
  for row in ws.iter_rows():
    for cell in row:
      cell.border = thin_border

  column_widths = {'B': 60, 'C': 45, 'D': 35, 'E': 35,'F': 70}
  for column, width in column_widths.items():
    ws.column_dimensions[column].width = width

  top_left_alignment = openpyxl.styles.Alignment(horizontal='left', vertical='top', wrap_text=True)
  for row in ws.iter_rows():
    for cell in row:
      cell.alignment = top_left_alignment
      new_height = max(cell.value.count('\n') + 1, 1) * 15
      if ws.row_dimensions[cell.row].height is None or ws.row_dimensions[cell.row].height < new_height:
        ws.row_dimensions[cell.row].height = new_height

  ws.freeze_panes = 'A2'
  wb.save(xlsx_file_path)